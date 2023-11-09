import openpyxl
from openpyxl.styles import PatternFill

def colorir_celas(sheet, linha, coluna_inicial, cor):
    # Loop pelas linhas de baixo para cima
    while linha >= 3:
        # Loop pelas colunas
        for coluna_atual in range(coluna_inicial, coluna_inicial + 5):
            # Colorir célula
            sheet.cell(row=linha, column=coluna_atual).fill = PatternFill(start_color=cor, end_color=cor, fill_type='solid')

        # Pular 2 colunas não coloridas
        coluna_inicial += 7

        # Próxima linha
        linha -= 1

# Carregar o arquivo do Excel
workbook = openpyxl.load_workbook('teste1.xlsx')  
sheet = workbook.active

# Definir as cores desejadas
cores = ['FFFF0000', 'FF00FF00', 'FF0000FF', 'FFFFFF00', 'FFFF00FF', 'FF00FFFF', 'FF800080', 'FF008000', 'FF000080', 'FFFFA500', 'FFA52A2A', 'FF8B0000', 'FF32CD32', 'FF0000CD', 'FF6A5ACD', 'FFA0522D', 'FF8B4513', 'FF2F4F4F']

linha_inicial = 41
coluna_inicial = 7  # Coluna G

# Loop para repetir o processo 18 vezes
i = 1
for i in range(18):
    cor_atual = cores[i % len(cores)]  # Selecionar cor com base na repetição

    # Chamar a função para colorir as células
    colorir_celas(sheet, linha_inicial, coluna_inicial, cor_atual)

    # Atualizar a linha inicial e coluna inicial para a próxima repetição
    linha_inicial = linha_inicial - 2
    coluna_inicial += 7
    
    workbook.save('teste1_pronto.xlsx')